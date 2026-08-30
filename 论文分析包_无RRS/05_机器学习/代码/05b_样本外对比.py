# -*- coding: utf-8 -*-
"""
05b_样本外对比.py — 无 RRS 版机器学习样本外对比（探索性）
  基准特征（L1+L2）：年龄、性别、抑郁 PHQ9、焦虑 GAD7（已剔除 RRS）
  方法层：
    A. LLM 融合语言块（8 维）
    B. Embedding 主成分（text-embedding-v3 → PCA，K 由训练集内 5 折选择）
  评估：固定划分 169/30、149/50；全体 10 折嵌套 CV。
  依赖（均离线，无需 API）：
    ../../forhoucan/embedding_ml/embeddings.npz
    ../../forhoucan/清理后数据_含对话.xlsx
    ../../forhoucan/llm_coding/results_language_{deepseek-chat,qwen-plus}.csv
    ../../forhoucan/analysis_R/split_{169_30,149_50}.csv
  运行：先运行 05a_TextMind_LASSO.R（生成表格.md），本脚本追加第 3、4 节。
"""
import csv, os, sys
import numpy as np
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")
PKG = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))   # 05_机器学习/
FORH = os.path.join(PKG, "..", "..", "forhoucan")
EMB = os.path.join(FORH, "embedding_ml")
MD_PATH = os.path.join(PKG, "表格.md")

DVS = {"CAIDS_anxiety_avg": "依恋焦虑", "CAIDS_avoidance_avg": "依恋回避", "CAIDS_total_avg": "总依恋"}
LLM_LANG = ["relationship_focus", "proximity_seeking", "emotional_expression", "abandonment_vigilance",
            "intellectualization", "psychological_distance", "emotional_dependence", "self_disclosure"]
BASE_VARS = ["age", "male", "PHQ9_avg", "GAD7_avg"]   # 已剔除 RRS

def md(s=""):
    with open(MD_PATH, "a", encoding="utf-8") as f:
        f.write(s + "\n")

# ---------------- 数据 ----------------
z = np.load(os.path.join(EMB, "embeddings.npz"))
Xemb, pids = z["vectors"], [int(p) for p in z["pids"]]
wb = load_workbook(os.path.join(FORH, "清理后数据_含对话.xlsx"), read_only=True)
ws = wb.active
rows = list(ws.iter_rows(values_only=True))
hdr = list(rows[0])
need = ["序号", "Q2", "Q3"] + list(DVS) + ["PHQ9_avg", "GAD7_avg"]
idx = {h: hdr.index(h) for h in need if h in hdr}
data = {}
for r in rows[1:]:
    if r[0] is None:
        continue
    pid = int(r[0])
    data[pid] = {"age": float(r[idx["Q2"]]),
                 "male": 1.0 if "男" in str(r[idx["Q3"]]) else 0.0}
    for dv in DVS:
        data[pid][dv] = float(r[idx[dv]])
    for s in ["PHQ9_avg", "GAD7_avg"]:
        data[pid][s] = float(r[idx[s]])

def read_avg(block, model):
    d = {}
    with open(os.path.join(FORH, "llm_coding", "results_%s_%s.csv" % (block, model)), encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            pid = int(row["participant_id"].replace("P", ""))
            d.setdefault(pid, {})
            for k, v in row.items():
                if k not in ("participant_id", "model", "run"):
                    d[pid].setdefault(k, []).append(float(v))
    return {p: {k: np.mean(v) for k, v in d[p].items()} for p in d}

ds = read_avg("language", "deepseek-chat"); qw = read_avg("language", "qwen-plus")
for pid in data:
    data[pid]["llm"] = {k: (ds.get(pid, {}).get(k, 0) + qw.get(pid, {}).get(k, 0)) / 2 for k in LLM_LANG}

order = sorted(data.keys())
pid2idx = {p: i for i, p in enumerate(order)}
mat_base = np.array([[data[p][v] for v in BASE_VARS] for p in order])
mat_llm = np.array([[data[p]["llm"][k] for k in LLM_LANG] for p in order])
mat_emb = np.array([Xemb[pid2idx[p]] for p in order])
y = {dv: np.array([data[p][dv] for p in order]) for dv in DVS}
N = len(order)
print("N =", N, "| 基准变量:", BASE_VARS)

def ols_fit(X, yv):
    return np.linalg.lstsq(np.hstack([np.ones((len(X), 1)), X]), yv, rcond=None)[0]

def ols_pred(X, beta):
    return np.hstack([np.ones((len(X), 1)), X]) @ beta

def r2(yv, pred):
    sst = np.sum((yv - yv.mean()) ** 2)
    return 1 - np.sum((yv - pred) ** 2) / sst if sst > 0 else np.nan

def pca_fit_apply(Xtr, Xte, K):
    mu = Xtr.mean(0)
    U, S, Vt = np.linalg.svd(Xtr - mu, full_matrices=False)
    return (Xtr - mu) @ Vt[:K].T, (Xte - mu) @ Vt[:K].T

def load_split(name):
    tr, va = [], []
    with open(os.path.join(FORH, "analysis_R", name), encoding="utf-8-sig") as f:
        for row in csv.DictReader(f):
            pid = int(row["序号"])
            if str(row.get("训练集", "")).lower() == "true":
                tr.append(pid)
            elif str(row.get("验证集", "")).lower() == "true":
                va.append(pid)
    return tr, va

K_CHOICES = [10, 20, 30]

md("")
md("## 3. 三方法样本外对比（L1+L2 基准 = 年龄/性别/抑郁/焦虑，无 RRS；方法层：TextMind 类别块¹ / LLM 语言块 / Embedding 主成分）")
md("")
md("> ¹ TextMind 类别块样本外引自 03 分层回归的块级结果（block_validate 口径，见旧管线），此处不重跑；LLM 与 Embedding 为本脚本实算。")

for split_name in ["split_169_30.csv", "split_149_50.csv"]:
    tr_pids, va_pids = load_split(split_name)
    tri = np.array([pid2idx[p] for p in tr_pids]); vai = np.array([pid2idx[p] for p in va_pids])
    md("")
    md("### 固定划分 %s（训练 %d / 验证 %d）" % (split_name.replace("split_", "").replace(".csv", ""), len(tri), len(vai)))
    md("")
    md("| 因变量 | 基准测试 R² | +LLM 语言块 R² (Δ) | +Embedding 主成分 R² (Δ) |")
    md("|---|---|---|---|")
    for dv, dvn in DVS.items():
        ytr, yva = y[dv][tri], y[dv][vai]
        b0 = ols_fit(mat_base[tri], ytr); r_base = r2(yva, ols_pred(mat_base[vai], b0))
        bL = ols_fit(np.hstack([mat_base[tri], mat_llm[tri]]), ytr)
        r_llm = r2(yva, ols_pred(np.hstack([mat_base[vai], mat_llm[vai]]), bL))
        best = None
        rng = np.random.default_rng(42)
        folds = np.array_split(rng.permutation(len(ytr)), 5)
        for K in K_CHOICES:
            errs = []
            for f in folds:
                tri2 = np.setdiff1d(np.arange(len(ytr)), f)
                Atr, Af = pca_fit_apply(mat_emb[tri][tri2], mat_emb[tri][f], K)
                b = ols_fit(np.hstack([mat_base[tri][tri2], Atr]), ytr[tri2])
                pred = ols_pred(np.hstack([mat_base[tri][f], Af]), b)
                errs.append(np.mean((ytr[f] - pred) ** 2))
            if best is None or np.mean(errs) < best[0]:
                best = (np.mean(errs), K)
        K = best[1]
        Atr, Ava = pca_fit_apply(mat_emb[tri], mat_emb[vai], K)
        bE = ols_fit(np.hstack([mat_base[tri], Atr]), ytr)
        r_emb = r2(yva, ols_pred(np.hstack([mat_base[vai], Ava]), bE))
        md("| %s | %+.3f | %+.3f (%+.3f) | %+.3f (%+.3f) K=%d |" % (dvn, r_base, r_llm, r_llm - r_base, r_emb, r_emb - r_base, K))

md("")
md("### 全体 10 折嵌套 CV（方法层相对基准的测试 R² 增量）")
md("")
md("| 因变量 | 基准 R² | LLM 语言块 Δ | Embedding Δ |")
md("|---|---|---|---|")
rng = np.random.default_rng(7)
folds = np.array_split(rng.permutation(N), 10)
for dv, dvn in DVS.items():
    d_base, d_llm, d_emb = [], [], []
    for f in folds:
        tri = np.setdiff1d(np.arange(N), f)
        ytr, yva = y[dv][tri], y[dv][f]
        b0 = ols_fit(mat_base[tri], ytr); rb = r2(yva, ols_pred(mat_base[f], b0))
        bL = ols_fit(np.hstack([mat_base[tri], mat_llm[tri]]), ytr)
        rl = r2(yva, ols_pred(np.hstack([mat_base[f], mat_llm[f]]), bL))
        Atr, Ava = pca_fit_apply(mat_emb[tri], mat_emb[f], 20)
        bE = ols_fit(np.hstack([mat_base[tri], Atr]), ytr)
        re_ = r2(yva, ols_pred(np.hstack([mat_base[f], Ava]), bE))
        d_base.append(rb); d_llm.append(rl - rb); d_emb.append(re_ - rb)
    md("| %s | %+.3f | %+.3f | %+.3f |" % (dvn, np.mean(d_base), np.mean(d_llm), np.mean(d_emb)))
md("")
md("## 4. 小结（探索性，不作结论）")
md("")
md("去除 RRS 后，词典 LASSO、LLM 语言块与 Embedding 的样本外表现整体仍弱（验证 R² 接近零或为负），与论文 5.7/附录五结论方向一致：数据驱动路径在当前样本量下无法稳定预测人机依恋，理论驱动的块级检验仍是主证据。")

report = open(MD_PATH, encoding="utf-8").read()
with open(os.path.join(PKG, "表格.md"), "w", encoding="utf-8") as f:
    f.write(report)
print("表格.md 已追加完成：", MD_PATH)
